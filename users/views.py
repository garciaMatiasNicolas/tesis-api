from django.shortcuts import render
from rest_framework.views import APIView
from rest_framework import viewsets
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import AllowAny
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser
from django.http import HttpResponse
from django.core.exceptions import ObjectDoesNotExist
from django.contrib.auth import get_user_model
from django.db.models import Q
from django.db import transaction
import pandas as pd
from io import BytesIO
from .serializer import UserSerializer, EmployeeSerializer, SupplierSerializer, EmployeeCreateSerializer, EmployeeUpdateSerializer, UserWithEmployeeSerializer
from .models import Employee, Supplier
from rest_framework.permissions import IsAuthenticated
from core.store.models import Branch
from core.crm.models import Customer
from .permissions import IsNotClientPermission

User = get_user_model()


class UserModelViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer

    def get_permissions(self):
        if self.action in ['create']:
            return [AllowAny()]
        elif self.action in ['list', 'retrieve', 'update', 'partial_update', 'destroy']:
            return [IsAuthenticated()]
        
        return super().get_permissions()
    
    ## Lista todos los usuarios
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)
    
    ## Lista un usuario por su ID
    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'])
    def me(self, request):
        """Obtener información del usuario actual incluyendo datos del empleado si corresponde"""
        # Usar el serializer extendido que incluye información del empleado
        serializer = UserWithEmployeeSerializer(request.user)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    ## Crea un nuevo usuario
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        
        if serializer.is_valid():
            user = serializer.save()
            return Response({"message": "user_created", "user": self.get_serializer(user).data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    ## Actualiza un usuario por su ID
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        
        if serializer.is_valid():
            serializer.save()
            return Response(
                {"message": "user_updated", "user": serializer.data}, 
                status=status.HTTP_200_OK
            )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


    ## Elimina un usuario por su ID
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return Response({"message": "user_deleted"}, status=status.HTTP_204_NO_CONTENT)


class EmailExistsAPIView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        email = request.GET.get('email')
        
        if not email:
            return Response(
                data={'error': 'Email parameter is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verificar si existe un usuario con este email
        user_exists = User.objects.filter(email=email).exists()
        
        if user_exists:
            # Ya existe un usuario con este email
            return Response(data={
                'available': False, 
                'exists': True,
                'has_user': True,
                'message': 'Ya existe una cuenta con este email. Por favor, inicia sesión.'
            }, status=status.HTTP_200_OK)
        
        # Verificar si existe un customer sin usuario con este email
        customer_without_user = Customer.objects.filter(
            email=email, 
            user__isnull=True
        ).exists()
        
        if customer_without_user:
            # Existe un customer sin usuario - puede registrarse y se vinculará automáticamente
            return Response(data={
                'available': True, 
                'exists': False,
                'has_user': False,
                'customer_exists': True,
                'message': 'Puedes crear tu cuenta. Vincularemos tu historial de compras previo.'
            }, status=status.HTTP_200_OK)
        
        # Email completamente disponible
        return Response(data={
            'available': True, 
            'exists': False,
            'has_user': False,
            'customer_exists': False,
            'message': 'Email disponible para registro.'
        }, status=status.HTTP_200_OK)


class VerifyIsClientAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        
        user_data = {
            'id': user.id,
            'email': user.email,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'role': user.role
        }
        
        # Verificar si existe un Customer asociado a este usuario
        is_customer = Customer.objects.filter(user=user).exists()
        
        if is_customer:
            return Response(data={'is_client': True, 'user': user_data}, status=status.HTTP_200_OK)
        else:
            return Response(data={'is_client': False, 'user': user_data}, status=status.HTTP_200_OK)


class EmployeeViewSet(viewsets.ModelViewSet):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer
    permission_classes = [IsNotClientPermission]

    def get_serializer_class(self):
        if self.action == 'create':
            return EmployeeCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return EmployeeUpdateSerializer
        return EmployeeSerializer

    def get_queryset(self):
        user = self.request.user
        
        if user.role == 'superadmin':
            # Superadmin puede ver todos los empleados
            return Employee.objects.all()
        elif user.role == 'manager':
            # Manager solo puede ver empleados de su sucursal
            try:
                manager_employee = Employee.objects.get(user=user)
                if manager_employee.branch:
                    return Employee.objects.filter(branch=manager_employee.branch)
                else:
                    # Si el manager no tiene sucursal asignada, no puede ver empleados
                    return Employee.objects.none()
            except Employee.DoesNotExist:
                # Si el manager no tiene registro de empleado, no puede ver empleados
                return Employee.objects.none()
        else:
            # Otros roles no pueden ver empleados
            return Employee.objects.none()

    def create(self, request, *args, **kwargs):
        user = request.user
        
        if user.role not in ['superadmin', 'manager']:
            return Response(
                {"error": "No tienes permisos para crear empleados"}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Si es manager, verificar que solo pueda crear empleados en su sucursal
        if user.role == 'manager':
            # Obtener la sucursal del manager
            try:
                manager_employee = Employee.objects.get(user=user)
                manager_branch = manager_employee.branch
                
                if not manager_branch:
                    return Response(
                        {"error": "No tienes una sucursal asignada para crear empleados"}, 
                        status=status.HTTP_403_FORBIDDEN
                    )
                
                # Verificar que la sucursal especificada sea la del manager
                branch_id = request.data.get('branch')
                if branch_id and int(branch_id) != manager_branch.id:
                    return Response(
                        {"error": "Solo puedes crear empleados en tu propia sucursal"}, 
                        status=status.HTTP_403_FORBIDDEN
                    )
                
                # Si no se especifica sucursal, asignar automáticamente la del manager
                if not branch_id:
                    request.data['branch'] = manager_branch.id
                    
            except Employee.DoesNotExist:
                return Response(
                    {"error": "No tienes registro de empleado para crear otros empleados"}, 
                    status=status.HTTP_403_FORBIDDEN
                )

        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(
                {"message": "Empleado creado exitosamente", "employee": serializer.data}, 
                status=status.HTTP_201_CREATED
            )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        user = request.user

        # Verificar permisos de edición
        if user.role == 'superadmin':
            # Superadmin puede editar cualquier empleado
            pass
        elif user.role == 'manager':
            # Manager solo puede editar empleados de su propia sucursal
            try:
                manager_employee = Employee.objects.get(user=user)
                manager_branch = manager_employee.branch
                
                if not manager_branch:
                    return Response(
                        {"error": "No tienes una sucursal asignada"}, 
                        status=status.HTTP_403_FORBIDDEN
                    )
                
                # Verificar que el empleado pertenezca a la sucursal del manager
                if not instance.branch or instance.branch.id != manager_branch.id:
                    return Response(
                        {"error": "Solo puedes editar empleados de tu propia sucursal"}, 
                        status=status.HTTP_403_FORBIDDEN
                    )
                    
            except Employee.DoesNotExist:
                return Response(
                    {"error": "No tienes registro de empleado"}, 
                    status=status.HTTP_403_FORBIDDEN
                )
        elif user.role == 'employee':
            # Employee solo puede editar su propio registro
            if instance.user != user:
                return Response(
                    {"error": "Solo puedes editar tus propios datos"}, 
                    status=status.HTTP_403_FORBIDDEN
                )
        else:
            return Response(
                {"error": "No tienes permisos para editar empleados"}, 
                status=status.HTTP_403_FORBIDDEN
            )

        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        if serializer.is_valid():
            serializer.save()
            return Response(
                {"message": "Empleado actualizado exitosamente", "employee": serializer.data}, 
                status=status.HTTP_200_OK
            )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        user = request.user

        # Solo superadmin y manager pueden eliminar empleados
        if user.role == 'superadmin':
            # Superadmin puede eliminar cualquier empleado
            pass
        elif user.role == 'manager':
            # Manager solo puede eliminar empleados de su propia sucursal
            try:
                manager_employee = Employee.objects.get(user=user)
                manager_branch = manager_employee.branch
                
                if not manager_branch:
                    return Response(
                        {"error": "No tienes una sucursal asignada"}, 
                        status=status.HTTP_403_FORBIDDEN
                    )
                
                # Verificar que el empleado pertenezca a la sucursal del manager
                if not instance.branch or instance.branch.id != manager_branch.id:
                    return Response(
                        {"error": "Solo puedes eliminar empleados de tu propia sucursal"}, 
                        status=status.HTTP_403_FORBIDDEN
                    )
                    
            except Employee.DoesNotExist:
                return Response(
                    {"error": "No tienes registro de empleado"}, 
                    status=status.HTTP_403_FORBIDDEN
                )
        else:
            return Response(
                {"error": "No tienes permisos para eliminar empleados"}, 
                status=status.HTTP_403_FORBIDDEN
            )

        User.objects.get(id=instance.user.id).delete()
        instance.delete()
        return Response(
            {"message": "Empleado eliminado exitosamente"}, 
            status=status.HTTP_204_NO_CONTENT
        )

    @action(detail=False, methods=['get'])
    def by_branch(self, request):
        """Obtener empleados por sucursal"""
        branch_id = request.query_params.get('branch_id')
        if not branch_id:
            return Response(
                {"error": "El parámetro branch_id es requerido"}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        user = request.user
        try:
            branch = Branch.objects.get(id=branch_id)
        except Branch.DoesNotExist:
            return Response(
                {"error": "La sucursal no existe"}, 
                status=status.HTTP_404_NOT_FOUND
            )

        # Verificar permisos
        if user.role == 'superadmin':
            # Superadmin puede ver empleados de cualquier sucursal
            pass
        elif user.role == 'manager':
            # Manager solo puede ver empleados de sus sucursales
            if branch.manager != user:
                return Response(
                    {"error": "No tienes permisos para ver empleados de esta sucursal"}, 
                    status=status.HTTP_403_FORBIDDEN
                )
        else:
            return Response(
                {"error": "No tienes permisos para esta operación"}, 
                status=status.HTTP_403_FORBIDDEN
            )

        employees = Employee.objects.filter(branch=branch)
        serializer = self.get_serializer(employees, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class SupplierViewSet(viewsets.ModelViewSet):
    queryset = Supplier.objects.all()
    serializer_class = SupplierSerializer
    permission_classes = [IsNotClientPermission]

    def list(self, request, *args, **kwargs):
        # Todos los usuarios autenticados pueden ver proveedores
        return super().list(request, *args, **kwargs)

    def retrieve(self, request, *args, **kwargs):
        # Todos los usuarios autenticados pueden ver un proveedor específico
        return super().retrieve(request, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        # Por ahora no hay restricciones para crear proveedores
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            supplier = serializer.save()
            return Response(
                {"message": "Proveedor creado exitosamente", "supplier": serializer.data}, 
                status=status.HTTP_201_CREATED
            )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        # Por ahora no hay restricciones para actualizar proveedores
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        
        if serializer.is_valid():
            serializer.save()
            return Response(
                {"message": "Proveedor actualizado exitosamente", "supplier": serializer.data}, 
                status=status.HTTP_200_OK
            )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        # Por ahora no hay restricciones para eliminar proveedores
        instance = self.get_object()
        instance.delete()
        return Response(
            {"message": "Proveedor eliminado exitosamente"}, 
            status=status.HTTP_204_NO_CONTENT
        )

    @action(detail=False, methods=['get'])
    def search(self, request):
        """Buscar proveedores por nombre o CUIT"""
        query = request.query_params.get('q', '')
        if not query:
            return Response(
                {"error": "El parámetro 'q' es requerido para la búsqueda"}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        suppliers = Supplier.objects.filter(
            Q(name__icontains=query) |
            Q(fantasy_name__icontains=query) |
            Q(cuit__icontains=query)
        )
        
        serializer = self.get_serializer(suppliers, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'])
    def export_template(self, request):
        """Exportar plantilla Excel vacía para importación de proveedores"""
        from openpyxl.styles import PatternFill, Font
        from openpyxl.utils import get_column_letter
        
        # Crear DataFrame con columnas en español
        template_data = {
            'Nombre': ['Proveedor Ejemplo S.A.'],
            'Nombre Fantasía': ['Proveedor Ejemplo'],
            'Email': ['contacto@proveedor.com'],
            'Teléfono': ['+54 11 1234-5678'],
            'Sitio Web': ['https://www.proveedor.com'],
            'CUIT': ['20-12345678-9'],
            'País': ['Argentina'],
            'Provincia/Estado': ['Buenos Aires'],
            'Ciudad': ['Buenos Aires'],
            'Código Postal': ['1000'],
            'Dirección': ['Av. Corrientes 1234'],
            'Días de Entrega': [7]
        }
        
        df = pd.DataFrame(template_data)
        
        # Crear archivo Excel en memoria
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Hoja 1: Plantilla de proveedores
            df.to_excel(writer, index=False, sheet_name='Proveedores')
            worksheet1 = writer.sheets['Proveedores']
            
            # Hoja 2: Instrucciones
            instructions = pd.DataFrame({
                'Instrucciones': [
                    '1. Complete los datos de los proveedores en la hoja "Proveedores"',
                    '2. Campo obligatorio: Nombre (debe ser único)',
                    '3. Campos opcionales: Nombre Fantasía, Email, Teléfono, Sitio Web, CUIT, País, Provincia/Estado, Ciudad, Código Postal, Dirección',
                    '4. Días de Entrega: días de entrega (default: 0)',
                    '5. No modifique los nombres de las columnas',
                    '6. Elimine esta fila de ejemplo antes de importar'
                ]
            })
            instructions.to_excel(writer, index=False, sheet_name='Instrucciones')
            worksheet2 = writer.sheets['Instrucciones']
            
            # Aplicar formato
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(color='FFFFFF')  # Sin bold
            
            # Formatear hoja Proveedores
            for cell in worksheet1[1]:
                cell.fill = header_fill
                cell.font = header_font
            for column in worksheet1.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet1.column_dimensions[column_letter].width = adjusted_width
            
            # Formatear hoja Instrucciones
            for cell in worksheet2[1]:
                cell.fill = header_fill
                cell.font = header_font
            for column in worksheet2.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 80)
                worksheet2.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        # Crear respuesta HTTP con archivo Excel
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_proveedores.xlsx"'
        return response
    
    @action(detail=False, methods=['get'])
    def export(self, request):
        """Exportar todos los proveedores a CSV"""
        suppliers = self.get_queryset()
        
        # Usar values() para obtener diccionarios directamente de la query
        data = suppliers.values(
            'id', 'name', 'fantasy_name', 'email', 'phone', 'website',
            'cuit', 'country', 'state', 'city', 'postal_code', 'address',
            'lead_time_days', 'created_at', 'updated_at'
        )
        
        # Crear DataFrame directamente desde los valores
        df = pd.DataFrame(list(data))
        
        # Renombrar columnas a español
        df.columns = [
            'ID', 'Nombre', 'Nombre Fantasía', 'Email', 'Teléfono', 'Sitio Web',
            'CUIT', 'País', 'Provincia/Estado', 'Ciudad', 'Código Postal', 'Dirección',
            'Días de Entrega', 'Fecha Creación', 'Fecha Actualización'
        ]
        
        # Formatear fechas
        if 'Fecha Creación' in df.columns:
            df['Fecha Creación'] = pd.to_datetime(df['Fecha Creación']).dt.strftime('%Y-%m-%d %H:%M:%S')
        if 'Fecha Actualización' in df.columns:
            df['Fecha Actualización'] = pd.to_datetime(df['Fecha Actualización']).dt.strftime('%Y-%m-%d %H:%M:%S')
        
        # Rellenar valores nulos con cadena vacía
        df = df.fillna('')
        
        # Crear archivo CSV en memoria con sep=; y decimal=,
        output = BytesIO()
        df.to_csv(output, index=False, encoding='utf-8-sig', sep=';', decimal=',')
        output.seek(0)
        
        # Crear respuesta HTTP con archivo CSV
        response = HttpResponse(
            output.getvalue(),
            content_type='text/csv; charset=utf-8-sig'
        )
        response['Content-Disposition'] = 'attachment; filename="proveedores_export.csv"'
        return response
    
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser])
    def import_data(self, request):
        """Importar proveedores desde archivo Excel (bulk create)"""
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No se proporcionó ningún archivo'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file = request.FILES['file']
        
        # Validar que sea un archivo Excel
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'El archivo debe ser un Excel (.xlsx o .xls)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Leer archivo Excel
            df = pd.read_excel(file, sheet_name='Proveedores')
            
            # Mapeo de columnas en español a nombres de campos del modelo
            column_mapping = {
                'Nombre': 'name',
                'Nombre Fantasía': 'fantasy_name',
                'Email': 'email',
                'Teléfono': 'phone',
                'Sitio Web': 'website',
                'CUIT': 'cuit',
                'País': 'country',
                'Provincia/Estado': 'state',
                'Ciudad': 'city',
                'Código Postal': 'postal_code',
                'Dirección': 'address',
                'Días de Entrega': 'lead_time_days'
            }
            
            # Renombrar columnas del DataFrame
            df = df.rename(columns=column_mapping)
            
            # Validar columna requerida
            if 'name' not in df.columns:
                return Response(
                    {'error': 'Falta la columna requerida: Nombre'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # ===== OPTIMIZACIÓN: Precargar proveedores existentes (1 query) =====
            existing_suppliers_qs = Supplier.objects.all()
            existing_suppliers_map = {s.name: s for s in existing_suppliers_qs}
            
            suppliers_to_create = []
            suppliers_to_update = []
            errors = []
            
            # ===== Procesar datos en memoria (0 queries adicionales) =====
            for index, row in df.iterrows():
                try:
                    # Validar nombre (obligatorio)
                    name = str(row['name']).strip() if pd.notna(row['name']) else None
                    if not name:
                        errors.append(f'Fila {index + 2}: Nombre es obligatorio')
                        continue
                    
                    # Verificar si ya existe un proveedor con este nombre (lookup O(1))
                    existing_supplier = existing_suppliers_map.get(name)
                    
                    # Preparar datos del proveedor
                    supplier_data = {
                        'name': name,
                        'fantasy_name': str(row['fantasy_name']).strip() if pd.notna(row.get('fantasy_name')) else None,
                        'email': str(row['email']).strip() if pd.notna(row.get('email')) else None,
                        'phone': str(row['phone']).strip() if pd.notna(row.get('phone')) else None,
                        'website': str(row['website']).strip() if pd.notna(row.get('website')) else None,
                        'cuit': str(row['cuit']).strip() if pd.notna(row.get('cuit')) else None,
                        'country': str(row['country']).strip() if pd.notna(row.get('country')) else None,
                        'state': str(row['state']).strip() if pd.notna(row.get('state')) else None,
                        'city': str(row['city']).strip() if pd.notna(row.get('city')) else None,
                        'postal_code': str(row['postal_code']).strip() if pd.notna(row.get('postal_code')) else None,
                        'address': str(row['address']).strip() if pd.notna(row.get('address')) else None,
                        'lead_time_days': int(row['lead_time_days']) if pd.notna(row.get('lead_time_days')) else 0
                    }
                    
                    if existing_supplier:
                        # Actualizar proveedor existente
                        for key, value in supplier_data.items():
                            setattr(existing_supplier, key, value)
                        suppliers_to_update.append(existing_supplier)
                    else:
                        # Crear nuevo proveedor
                        supplier = Supplier(**supplier_data)
                        suppliers_to_create.append(supplier)
                    
                except Exception as e:
                    errors.append(f'Fila {index + 2}: {str(e)}')
                    continue
            
            # ===== Ejecutar operaciones en bulk (2 queries totales) =====
            with transaction.atomic():
                # Crear proveedores nuevos (1 query)
                if suppliers_to_create:
                    Supplier.objects.bulk_create(suppliers_to_create)
                
                # Actualizar proveedores existentes (1 query)
                if suppliers_to_update:
                    Supplier.objects.bulk_update(
                        suppliers_to_update,
                        ['name', 'fantasy_name', 'email', 'phone', 'website', 'cuit',
                         'country', 'state', 'city', 'postal_code', 'address', 'lead_time_days']
                    )
            
            return Response({
                'message': f'Importación completada. Creados: {len(suppliers_to_create)}, Actualizados: {len(suppliers_to_update)}',
                'created': len(suppliers_to_create),
                'updated': len(suppliers_to_update),
                'errors': errors if errors else []
            }, status=status.HTTP_200_OK if not errors else status.HTTP_207_MULTI_STATUS)
            
        except Exception as e:
            return Response(
                {'error': f'Error al procesar el archivo: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
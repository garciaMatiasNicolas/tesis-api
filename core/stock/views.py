from django.shortcuts import render
from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination
from django.http import HttpResponse
from django.utils import timezone
import os
from decimal import Decimal
from django.conf import settings
import pandas as pd
import openpyxl
from io import BytesIO
from .models import Product, Category, Subcategory, Warehouse, ProductUnit, Stock, StockMovement
from .serializers import ProductSerializer, CategorySerializer, SubcategorySerializer, WarehouseSerializer, ProductUnitSerializer, StockSerializer, StockMovementSerializer
from users.permissions import IsNotClientPermission
from core.store.models import Store, Branch
from django.db import transaction


class ProductPagination(PageNumberPagination):
    page_size = 5
    page_size_query_param = 'page_size'
    max_page_size = 100


class StockPagination(PageNumberPagination):
    page_size = 5
    page_size_query_param = 'page_size'
    max_page_size = 100


class StockMovementPagination(PageNumberPagination):
    page_size = 5
    page_size_query_param = 'page_size'
    max_page_size = 100


class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes = [IsNotClientPermission]
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    pagination_class = ProductPagination

    def list(self, request, *args, **kwargs):
        """
        Listar todos los productos con filtros opcionales y paginación.
        Query params:
        - search: Búsqueda por SKU, descripción o nombre de categoría
        - category: ID de la categoría
        - status: Estado del producto ('active' o 'discontinued')
        - page: Número de página (default: 1)
        - page_size: Tamaño de página (default: 5, max: 100)
        """
        from django.db.models import Q
        queryset = self.get_queryset()
        
        # Filtros opcionales
        search_term = request.query_params.get('search', None)
        category_id = request.query_params.get('category', None)
        status_filter = request.query_params.get('status', None)
        supplier = request.query_params.get('supplier', None)
        all_products = request.query_params.get('all', None)
        
        if search_term:
            queryset = queryset.filter(
                Q(sku__icontains=search_term) | 
                Q(description__icontains=search_term) |
                Q(category__name__icontains=search_term)
            )
        
        if category_id:
            queryset = queryset.filter(category_id=category_id)
        
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        if supplier:
            queryset = queryset.filter(supplier_id=supplier)
        
        # Aplicar paginación
        if not all_products:
            page = self.paginate_queryset(queryset)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)
        
        # Si no hay paginación (no debería ocurrir), devolver todos
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)
    
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
    def perform_create(self, serializer):
        # Create the product
        product = serializer.save()
        
        # Create a base stock record with quantity 0 and no warehouse/branch
        Stock.objects.create(
            product=product,
            quantity=0.0000,
            warehouse=None,
            branch=None
        )
    
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    def destroy(self, request, *args, **kwargs):
        """Eliminación lógica: marca el producto como descontinuado"""
        instance = self.get_object()
        instance.status = 'discontinued'
        instance.save()
        return Response({
            'message': 'Producto descontinuado exitosamente',
            'status': 'discontinued'
        }, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['delete'])
    def permanent_delete(self, request, pk=None):
        """Eliminación física permanente del producto"""
        product = self.get_object()
        product.delete()
        return Response({
            'message': 'Producto eliminado permanentemente'
        }, status=status.HTTP_204_NO_CONTENT)
    
    @action(detail=True, methods=['post'])
    def reactivate(self, request, pk=None):
        """Reactivar un producto descontinuado"""
        product = self.get_object()
        if product.status != 'discontinued':
            return Response(
                {'error': 'El producto ya está activo'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        product.status = 'active'
        product.save()
        return Response({
            'message': 'Producto reactivado exitosamente',
            'status': 'active'
        }, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def upload_image(self, request, pk=None):
        """
        Endpoint para subir imágenes de producto al servidor local.
        Soporta hasta 3 imágenes por producto.
        Estructura: media/images/{store}/{sku}/image_1.jpg
        """
        product = self.get_object()
        
        # Determinar qué imagen se está subiendo (image_1, image_2, o image_3)
        image_slot = request.data.get('slot', 'image_1')  # Por defecto image_1
        if image_slot not in ['image_1', 'image_2', 'image_3']:
            return Response(
                {'error': 'Slot de imagen inválido. Use: image_1, image_2, o image_3'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if 'image' not in request.FILES:
            return Response(
                {'error': 'No se encontró archivo de imagen'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            image_file = request.FILES['image']
            
            # Validar tamaño (máximo 5MB)
            if image_file.size > 5 * 1024 * 1024:
                return Response(
                    {'error': 'El archivo de imagen no puede ser mayor a 5MB'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Validar tipo
            allowed_types = ['image/jpeg', 'image/jpg', 'image/png', 'image/gif', 'image/webp']
            if image_file.content_type not in allowed_types:
                return Response(
                    {'error': 'Solo se permiten archivos JPEG, PNG, GIF y WebP'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Eliminar imagen anterior si existe
            current_image = getattr(product, image_slot)
            if current_image:
                try:
                    # Construir ruta completa del archivo
                    old_image_path = os.path.join(settings.MEDIA_ROOT, str(current_image))
                    if os.path.exists(old_image_path):
                        os.remove(old_image_path)
                except Exception as e:
                    print(f"Error al eliminar imagen anterior: {e}")
            
            # Guardar nueva imagen usando el ImageField
            setattr(product, image_slot, image_file)
            product.save()
            
            # Obtener URL completa de la imagen
            image_url = request.build_absolute_uri(getattr(product, image_slot).url)
            
            return Response({
                'message': f'Imagen {image_slot} subida exitosamente',
                'slot': image_slot,
                'image_url': image_url,
                'all_images': {
                    'image_1': request.build_absolute_uri(product.image_1.url) if product.image_1 else None,
                    'image_2': request.build_absolute_uri(product.image_2.url) if product.image_2 else None,
                    'image_3': request.build_absolute_uri(product.image_3.url) if product.image_3 else None,
                }
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {'error': f'Error al subir imagen: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['delete'])
    def delete_image(self, request, pk=None):
        """
        Endpoint para eliminar imagen de producto del servidor local.
        Puede eliminar una imagen específica o todas.
        """
        product = self.get_object()
        
        # Determinar qué imagen eliminar
        image_slot = request.query_params.get('slot', 'all')  # Por defecto elimina todas
        
        if image_slot not in ['image_1', 'image_2', 'image_3', 'all']:
            return Response(
                {'error': 'Slot de imagen inválido. Use: image_1, image_2, image_3, o all'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            deleted_images = []
            
            # Eliminar imagen específica o todas
            slots_to_delete = ['image_1', 'image_2', 'image_3'] if image_slot == 'all' else [image_slot]
            
            for slot in slots_to_delete:
                current_image = getattr(product, slot)
                
                if current_image:
                    try:
                        # Construir ruta completa del archivo
                        image_path = os.path.join(settings.MEDIA_ROOT, str(current_image))
                        # Eliminar archivo físico si existe
                        if os.path.exists(image_path):
                            os.remove(image_path)
                            deleted_images.append(slot)
                        
                        # Limpiar campo en base de datos
                        setattr(product, slot, None)
                    except Exception as e:
                        print(f"Error al eliminar {slot}: {e}")
            
            if not deleted_images:
                return Response(
                    {'error': 'No hay imágenes para eliminar'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            product.save()
            
            return Response({
                'message': f'Imagen(es) eliminada(s) exitosamente: {", ".join(deleted_images)}',
                'deleted_slots': deleted_images,
                'remaining_images': {
                    'image_1': request.build_absolute_uri(product.image_1.url) if product.image_1 else None,
                    'image_2': request.build_absolute_uri(product.image_2.url) if product.image_2 else None,
                    'image_3': request.build_absolute_uri(product.image_3.url) if product.image_3 else None,
                }
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {'error': f'Error al eliminar imagen: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def export_template(self, request):
        """Exportar plantilla Excel vacía para importación de productos"""
        from openpyxl.styles import PatternFill, Font
        from openpyxl.utils import get_column_letter
        
        # Crear DataFrame con columnas en español
        template_data = {
            'SKU': ['PROD-001'],
            'Descripción': ['Producto de ejemplo'],
            'Precio': [100.00],
            'Precio Costo': [50.00],
            'Stock Seguridad': [10.0000],
            'Tipo Unidad': ['count'],
            'Unidad Base': ['unit'],
            'ID Categoría': [1],
            'ID Subcategoría': [1],
            'ID Proveedor': [1],
            'Estado': ['active']
        }
        
        df_template = pd.DataFrame(template_data)
        
        # Obtener catálogos de referencia
        categories = Category.objects.all().values('id', 'name')
        df_categories = pd.DataFrame(list(categories))
        if not df_categories.empty:
            df_categories.columns = ['ID', 'Nombre']
        else:
            df_categories = pd.DataFrame({'ID': [], 'Nombre': []})
        
        subcategories = Subcategory.objects.all().values('id', 'name', 'category__name')
        df_subcategories = pd.DataFrame(list(subcategories))
        if not df_subcategories.empty:
            df_subcategories.columns = ['ID', 'Nombre', 'Categoría']
        else:
            df_subcategories = pd.DataFrame({'ID': [], 'Nombre': [], 'Categoría': []})
        
        from users.models import Supplier
        suppliers = Supplier.objects.all().values('id', 'name')
        df_suppliers = pd.DataFrame(list(suppliers))
        if not df_suppliers.empty:
            df_suppliers.columns = ['ID', 'Nombre']
        else:
            df_suppliers = pd.DataFrame({'ID': [], 'Nombre': []})
        
        # Crear archivo Excel en memoria
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Hoja 1: Plantilla de productos
            df_template.to_excel(writer, index=False, sheet_name='Productos')
            worksheet1 = writer.sheets['Productos']
            
            # Hoja 2: Catálogo de Categorías
            df_categories.to_excel(writer, index=False, sheet_name='Categorías')
            worksheet2 = writer.sheets['Categorías']
            
            # Hoja 3: Catálogo de Subcategorías
            df_subcategories.to_excel(writer, index=False, sheet_name='Subcategorías')
            worksheet3 = writer.sheets['Subcategorías']
            
            # Hoja 4: Catálogo de Proveedores
            df_suppliers.to_excel(writer, index=False, sheet_name='Proveedores')
            worksheet4 = writer.sheets['Proveedores']
            
            # Hoja 5: Instrucciones
            instructions = pd.DataFrame({
                'Instrucciones': [
                    '1. Complete los datos de los productos en la hoja "Productos"',
                    '2. Para "ID Categoría", consulte la hoja "Categorías" con los IDs disponibles',
                    '3. Para "ID Subcategoría", consulte la hoja "Subcategorías" con los IDs disponibles',
                    '4. Para "ID Proveedor", consulte la hoja "Proveedores" con los IDs disponibles',
                    '5. Tipo Unidad: count, weight, volume',
                    '6. Unidad Base: unit, kg, g, l, ml',
                    '7. Estado: active o discontinued',
                    '8. Campos obligatorios: SKU, Descripción, Precio, Precio Costo',
                    '9. No modifique los nombres de las columnas',
                    '10. Elimine esta fila de ejemplo antes de importar'
                ]
            })
            instructions.to_excel(writer, index=False, sheet_name='Instrucciones')
            worksheet5 = writer.sheets['Instrucciones']
            
            # Aplicar formato a todas las hojas
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(color='FFFFFF')
            
            for worksheet in [worksheet1, worksheet2, worksheet3, worksheet4, worksheet5]:
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 80 if worksheet == worksheet5 else 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_productos.xlsx"'
        return response
    
    @action(detail=False, methods=['get'])
    def export(self, request):
        """Exportar todos los productos a CSV"""
        products = self.get_queryset()
        
        # Usar values() para obtener diccionarios directamente de la query
        data = products.values(
            'id', 'sku', 'description', 'price', 'cost_price', 'safety_stock',
            'unit_type', 'base_unit_name', 'status',
            'category__id', 'category__name',
            'subcategory__id', 'subcategory__name',
            'supplier__id', 'supplier__name',
            'created_at', 'updated_at'
        )
        
        df = pd.DataFrame(list(data))
        
        # Renombrar columnas a español
        df.columns = [
            'ID', 'SKU', 'Descripción', 'Precio', 'Precio Costo', 'Stock Seguridad',
            'Tipo Unidad', 'Unidad Base', 'Estado',
            'ID Categoría', 'Nombre Categoría',
            'ID Subcategoría', 'Nombre Subcategoría',
            'ID Proveedor', 'Nombre Proveedor',
            'Fecha Creación', 'Fecha Actualización'
        ]
        
        # Formatear fechas
        if 'Fecha Creación' in df.columns:
            df['Fecha Creación'] = pd.to_datetime(df['Fecha Creación']).dt.strftime('%Y-%m-%d %H:%M:%S')
        if 'Fecha Actualización' in df.columns:
            df['Fecha Actualización'] = pd.to_datetime(df['Fecha Actualización']).dt.strftime('%Y-%m-%d %H:%M:%S')
        
        df = df.fillna('')
        
        # Crear archivo CSV en memoria
        output = BytesIO()
        df.to_csv(output, index=False, encoding='utf-8-sig', sep=';', decimal=',')
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='text/csv; charset=utf-8-sig'
        )
        response['Content-Disposition'] = 'attachment; filename="productos_export.csv"'
        return response
    
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser])
    def import_data(self, request):
        """Importar productos desde archivo Excel"""
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No se encontró el archivo'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file = request.FILES['file']
        
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'El archivo debe ser un Excel (.xlsx o .xls)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Leer archivo Excel
            df = pd.read_excel(file, sheet_name='Productos')
            
            # Mapeo de columnas
            column_mapping = {
                'SKU': 'sku',
                'Descripción': 'description',
                'Precio': 'price',
                'Precio Costo': 'cost_price',
                'Stock Seguridad': 'safety_stock',
                'Tipo Unidad': 'unit_type',
                'Unidad Base': 'base_unit_name',
                'ID Categoría': 'category_id',
                'ID Subcategoría': 'subcategory_id',
                'ID Proveedor': 'supplier_id',
                'Estado': 'status'
            }
            
            df = df.rename(columns=column_mapping)
            
            # Validar columnas requeridas
            required_columns = ['sku', 'description', 'price', 'cost_price']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                return Response(
                    {'error': f'Faltan columnas requeridas: {", ".join(missing_columns)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # ===== OPTIMIZACIÓN: Precargar TODOS los datos necesarios (solo 4 queries) =====
            from users.models import Supplier
            
            # 1. Precargar productos existentes indexados por SKU (1 query)
            existing_products_qs = Product.objects.all()
            existing_products_map = {p.sku: p for p in existing_products_qs}
            
            # 2. Precargar categorías indexadas por ID (1 query)
            categories_qs = Category.objects.all()
            categories_map = {c.id: c for c in categories_qs}
            
            # 3. Precargar subcategorías indexadas por ID (1 query)
            subcategories_qs = Subcategory.objects.all()
            subcategories_map = {s.id: s for s in subcategories_qs}
            
            # 4. Precargar proveedores indexados por ID (1 query)
            suppliers_qs = Supplier.objects.all()
            suppliers_map = {s.id: s for s in suppliers_qs}
            
            products_to_create = []
            products_to_update = []
            errors = []
            
            # ===== Procesar datos en memoria (0 queries adicionales) =====
            for index, row in df.iterrows():
                try:
                    # Validar campos requeridos
                    if pd.isna(row['sku']) or pd.isna(row['description']):
                        errors.append(f"Fila {index + 2}: SKU y Descripción son obligatorios")
                        continue
                    
                    sku = str(row['sku']).strip()
                    
                    # Buscar producto existente (lookup O(1) en diccionario)
                    existing_product = existing_products_map.get(sku)
                    
                    # Obtener relaciones FK usando diccionarios (lookup O(1))
                    category = None
                    if pd.notna(row.get('category_id')):
                        category_id = int(row['category_id'])
                        category = categories_map.get(category_id)
                        if not category:
                            errors.append(f"Fila {index + 2}: Categoría con ID {category_id} no existe")
                            continue
                    
                    subcategory = None
                    if pd.notna(row.get('subcategory_id')):
                        subcategory_id = int(row['subcategory_id'])
                        subcategory = subcategories_map.get(subcategory_id)
                        if not subcategory:
                            errors.append(f"Fila {index + 2}: Subcategoría con ID {subcategory_id} no existe")
                            continue
                    
                    supplier = None
                    if pd.notna(row.get('supplier_id')):
                        supplier_id = int(row['supplier_id'])
                        supplier = suppliers_map.get(supplier_id)
                        if not supplier:
                            errors.append(f"Fila {index + 2}: Proveedor con ID {supplier_id} no existe")
                            continue
                    
                    # Preparar datos del producto
                    product_data = {
                        'sku': sku,
                        'description': str(row['description']).strip(),
                        'price': float(row['price']) if pd.notna(row['price']) else 0.0,
                        'cost_price': float(row['cost_price']) if pd.notna(row['cost_price']) else 0.0,
                        'safety_stock': float(row.get('safety_stock', 0.0)) if pd.notna(row.get('safety_stock')) else 0.0,
                        'unit_type': str(row.get('unit_type', 'count')).strip() if pd.notna(row.get('unit_type')) else 'count',
                        'base_unit_name': str(row.get('base_unit_name', 'unit')).strip() if pd.notna(row.get('base_unit_name')) else 'unit',
                        'status': str(row.get('status', 'active')).strip() if pd.notna(row.get('status')) else 'active',
                        'category': category,
                        'subcategory': subcategory,
                        'supplier': supplier
                    }
                    
                    if existing_product:
                        # Actualizar producto existente
                        for key, value in product_data.items():
                            setattr(existing_product, key, value)
                        products_to_update.append(existing_product)
                    else:
                        # Crear nuevo producto
                        product = Product(**product_data)
                        products_to_create.append(product)
                
                except Exception as e:
                    errors.append(f"Fila {index + 2}: {str(e)}")
                    continue
            
            # ===== Ejecutar operaciones en bulk (2-3 queries totales) =====
            with transaction.atomic():
                # Crear productos nuevos (1 query)
                if products_to_create:
                    created_products = Product.objects.bulk_create(products_to_create)
                    # Crear stock base para cada producto nuevo (1 query)
                    stock_records = [
                        Stock(product=product, quantity=0.0000, warehouse=None, branch=None)
                        for product in created_products
                    ]
                    Stock.objects.bulk_create(stock_records)
                
                # Actualizar productos existentes (1 query)
                if products_to_update:
                    Product.objects.bulk_update(
                        products_to_update,
                        ['description', 'price', 'cost_price', 'safety_stock', 'unit_type', 
                         'base_unit_name', 'status', 'category', 'subcategory', 'supplier']
                    )
            
            return Response({
                'message': f'Importación completada. Creados: {len(products_to_create)}, Actualizados: {len(products_to_update)}',
                'created': len(products_to_create),
                'updated': len(products_to_update),
                'errors': errors if errors else None
            }, status=status.HTTP_200_OK if not errors else status.HTTP_207_MULTI_STATUS)
        
        except Exception as e:
            return Response(
                {'error': f'Error al procesar el archivo: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [IsNotClientPermission]

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)
    
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
    def update(self, request, *args, **kwargs): 
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)


class SubcategoryViewSet(viewsets.ModelViewSet):
    queryset = Subcategory.objects.all()
    serializer_class = SubcategorySerializer
    permission_classes = [IsNotClientPermission]

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)
    
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)


class WarehouseViewSet(viewsets.ModelViewSet):
    queryset = Warehouse.objects.all()
    serializer_class = WarehouseSerializer
    permission_classes = [IsNotClientPermission]

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)
    
    def create(self, request, *args, **kwargs):
        if request.user.role == 'superadmin':
            request.data['store'] = Store.objects.first().id  # Asignar al primer store disponible
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        
        else:
            return Response(
                {'error': 'Solo superadministradores pueden crear depósitos'},
                status=status.HTTP_403_FORBIDDEN
            )
    
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        # Verificar si el depósito tiene stock antes de eliminar
        if instance.stocks.exists():
            return Response(
                {'error': 'No se puede eliminar el depósito porque tiene stock asociado'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        self.perform_destroy(instance)
        return Response({
            'message': 'Depósito eliminado exitosamente'
        }, status=status.HTTP_204_NO_CONTENT)
    
    @action(detail=True, methods=['get'])
    def stock(self, request, pk=None):
        """Obtener el stock del depósito"""
        warehouse = self.get_object()
        stocks = warehouse.stocks.select_related('product').all()
        
        stock_data = [{
            'product_id': stock.product.id,
            'product_sku': stock.product.sku,
            'product_description': stock.product.description,
            'quantity': stock.quantity,
            'updated_at': stock.updated_at
        } for stock in stocks]
        
        return Response({
            'warehouse': WarehouseSerializer(warehouse).data,
            'stock': stock_data,
            'total_items': len(stock_data)
        }, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'])
    def export_template(self, request):
        """Exportar plantilla Excel vacía para importación de depósitos"""
        from openpyxl.styles import PatternFill, Font
        from openpyxl.utils import get_column_letter
        
        # Crear DataFrame con columnas en español
        template_data = {
            'Nombre': ['Depósito Central'],
            'Dirección': ['Av. Principal 123'],
            'Ciudad': ['Buenos Aires'],
            'Provincia/Estado': ['Buenos Aires'],
            'País': ['Argentina'],
            'ID Tienda': [1]  # ID de la tienda
        }
        
        df_template = pd.DataFrame(template_data)
        
        # Obtener catálogo de tiendas para referencia
        stores = Store.objects.all().values('id', 'name')
        df_stores = pd.DataFrame(list(stores))
        if not df_stores.empty:
            df_stores.columns = ['ID', 'Nombre']
        else:
            df_stores = pd.DataFrame({'ID': [], 'Nombre': []})
        
        # Crear archivo Excel en memoria
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Hoja 1: Plantilla de depósitos
            df_template.to_excel(writer, index=False, sheet_name='Depósitos')
            worksheet1 = writer.sheets['Depósitos']
            
            # Hoja 2: Catálogo de Tiendas (referencia para ID Tienda)
            df_stores.to_excel(writer, index=False, sheet_name='Tiendas')
            worksheet2 = writer.sheets['Tiendas']
            
            # Hoja 3: Instrucciones
            instructions = pd.DataFrame({
                'Instrucciones': [
                    '1. Complete los datos de los depósitos en la hoja "Depósitos"',
                    '2. Para el campo "ID Tienda", consulte la hoja "Tiendas" con los IDs disponibles',
                    '3. Campos obligatorios: Nombre, Dirección, Ciudad, Provincia/Estado, País, ID Tienda',
                    '4. No modifique los nombres de las columnas',
                    '5. Elimine esta fila de ejemplo antes de importar'
                ]
            })
            instructions.to_excel(writer, index=False, sheet_name='Instrucciones')
            worksheet3 = writer.sheets['Instrucciones']
            
            # Aplicar formato a cada hoja
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(color='FFFFFF')  # Sin bold como solicitó el usuario
            
            # Formatear hoja Depósitos
            for cell in worksheet1[1]:
                cell.fill = header_fill
                cell.font = header_font
            for column in worksheet1.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet1.column_dimensions[column_letter].width = adjusted_width
            
            # Formatear hoja Tiendas
            for cell in worksheet2[1]:
                cell.fill = header_fill
                cell.font = header_font
            for column in worksheet2.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet2.column_dimensions[column_letter].width = adjusted_width
            
            # Formatear hoja Instrucciones
            for cell in worksheet3[1]:
                cell.fill = header_fill
                cell.font = header_font
            for column in worksheet3.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 80)  # Instrucciones pueden ser más anchas
                worksheet3.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        # Crear respuesta HTTP con archivo Excel
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_depositos.xlsx"'
        return response
    
    @action(detail=False, methods=['get'])
    def export(self, request):
        """Exportar todos los depósitos a CSV"""
        warehouses = self.get_queryset()
        
        # Usar values() para obtener diccionarios directamente de la query
        data = warehouses.values(
            'id', 'name', 'address', 'city', 'state', 'country', 
            'store__id', 'store__name', 'created_at'
        )
        
        # Crear DataFrame directamente desde los valores
        df = pd.DataFrame(list(data))
        
        # Renombrar columnas a español
        df.columns = [
            'ID', 'Nombre', 'Dirección', 'Ciudad', 'Provincia/Estado', 
            'País', 'ID Tienda', 'Nombre Tienda', 'Fecha Creación'
        ]
        
        # Formatear fecha
        if 'Fecha Creación' in df.columns:
            df['Fecha Creación'] = pd.to_datetime(df['Fecha Creación']).dt.strftime('%Y-%m-%d %H:%M:%S')
        
        # Rellenar valores nulos con cadena vacía
        df = df.fillna('')
        
        # Crear archivo CSV en memoria con sep=; y decimal=,
        output = BytesIO()
        df.to_csv(output, index=False, encoding='utf-8-sig', sep=';', decimal=',')
        output.seek(0)
        
        # Crear respuesta HTTP con archivo CSV
        response = HttpResponse(
            output.getvalue(),
            content_type='text/csv; charset=utf-8-sig'
        )
        response['Content-Disposition'] = 'attachment; filename="depositos_export.csv"'
        return response
    
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser])
    def import_data(self, request):
        """Importar depósitos desde archivo Excel (bulk create)"""
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No se proporcionó ningún archivo'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file = request.FILES['file']
        
        # Validar que sea un archivo Excel
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'El archivo debe ser un Excel (.xlsx o .xls)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Leer archivo Excel
            df = pd.read_excel(file, sheet_name='Depósitos')
            
            # Mapeo de columnas en español a nombres de campos del modelo
            column_mapping = {
                'Nombre': 'name',
                'Dirección': 'address',
                'Ciudad': 'city',
                'Provincia/Estado': 'state',
                'País': 'country',
                'ID Tienda': 'store_id'
            }
            
            # Renombrar columnas del DataFrame
            df = df.rename(columns=column_mapping)
            
            # Validar columnas requeridas
            required_columns = ['name', 'address', 'city', 'state', 'country', 'store_id']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                # Traducir nombres de columnas faltantes de vuelta a español para el mensaje de error
                reverse_mapping = {v: k for k, v in column_mapping.items()}
                missing_spanish = [reverse_mapping.get(col, col) for col in missing_columns]
                return Response(
                    {'error': f'Faltan columnas requeridas: {", ".join(missing_spanish)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # ===== OPTIMIZACIÓN: Precargar datos (2 queries) =====
            # 1. Precargar tiendas indexadas por ID (1 query)
            stores_qs = Store.objects.all()
            stores_map = {s.id: s for s in stores_qs}
            
            # 2. Precargar depósitos existentes con clave compuesta (nombre, store_id) (1 query)
            existing_warehouses_qs = Warehouse.objects.select_related('store').all()
            existing_warehouses_map = {(w.name, w.store_id): w for w in existing_warehouses_qs}
            
            warehouses_to_create = []
            warehouses_to_update = []
            errors = []
            
            # ===== Procesar datos en memoria (0 queries adicionales) =====
            for index, row in df.iterrows():
                try:
                    # Validar campos requeridos
                    if pd.isna(row['name']):
                        errors.append(f'Fila {index + 2}: Nombre es obligatorio')
                        continue
                    
                    # Validar que la tienda existe
                    store_id = int(row['store_id']) if pd.notna(row['store_id']) else None
                    if not store_id:
                        errors.append(f'Fila {index + 2}: ID Tienda es obligatorio')
                        continue
                    
                    # Buscar tienda en diccionario (lookup O(1))
                    store = stores_map.get(store_id)
                    if not store:
                        errors.append(f'Fila {index + 2}: Tienda con ID {store_id} no existe')
                        continue
                    
                    # Verificar si el depósito ya existe usando clave compuesta (lookup O(1))
                    name = str(row['name']).strip()
                    existing_warehouse = existing_warehouses_map.get((name, store_id))
                    
                    # Preparar datos del depósito
                    warehouse_data = {
                        'name': name,
                        'address': str(row['address']).strip() if pd.notna(row['address']) else '',
                        'city': str(row['city']).strip() if pd.notna(row['city']) else '',
                        'state': str(row['state']).strip() if pd.notna(row['state']) else '',
                        'country': str(row['country']).strip() if pd.notna(row['country']) else '',
                        'store': store
                    }
                    
                    if existing_warehouse:
                        # Actualizar depósito existente
                        for key, value in warehouse_data.items():
                            setattr(existing_warehouse, key, value)
                        warehouses_to_update.append(existing_warehouse)
                    else:
                        # Crear nuevo depósito
                        warehouse = Warehouse(**warehouse_data)
                        warehouses_to_create.append(warehouse)
                    
                except Exception as e:
                    errors.append(f'Fila {index + 2}: {str(e)}')
                    continue
            
            # ===== Ejecutar operaciones en bulk (2 queries totales) =====
            with transaction.atomic():
                # Crear depósitos nuevos (1 query)
                if warehouses_to_create:
                    Warehouse.objects.bulk_create(warehouses_to_create)
                
                # Actualizar depósitos existentes (1 query)
                if warehouses_to_update:
                    Warehouse.objects.bulk_update(
                        warehouses_to_update,
                        ['name', 'address', 'city', 'state', 'country', 'store']
                    )
            
            return Response({
                'message': f'Importación completada. Creados: {len(warehouses_to_create)}, Actualizados: {len(warehouses_to_update)}',
                'created': len(warehouses_to_create),
                'updated': len(warehouses_to_update),
                'errors': errors if errors else []
            }, status=status.HTTP_200_OK if not errors else status.HTTP_207_MULTI_STATUS)
            
        except Exception as e:
            return Response(
                {'error': f'Error al procesar el archivo: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ProductUnitViewSet(viewsets.ModelViewSet):
    queryset = ProductUnit.objects.all()
    serializer_class = ProductUnitSerializer
    permission_classes = [IsNotClientPermission]

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        
        # Filtrar por producto si se proporciona el parámetro 'product'
        product_id = request.query_params.get('product', None)
        
        if product_id is not None:
            prod = Product.objects.get(id=product_id)  # Verificar que el producto exista
            queryset = queryset.filter(product=prod)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)
    
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)
    
    @action(detail=False, methods=['get'])
    def export_template(self, request):
        """Exportar plantilla Excel para importación de unidades de producto"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        output = BytesIO()
        workbook = Workbook()
        
        # ===== Sheet 1: Plantilla de unidades =====
        worksheet1 = workbook.active
        worksheet1.title = 'Unidades de Producto'
        
        # Headers
        headers = ['SKU', 'Nombre', 'Conversion Factor']
        worksheet1.append(headers)
        
        # Ejemplo
        worksheet1.append(['PROD-001', 'Caja x12', 12.0000])
        
        # Formato de headers
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF')
        for cell in worksheet1[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Ajustar anchos
        for column in worksheet1.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet1.column_dimensions[column_letter].width = adjusted_width
        
        # ===== Sheet 2: Catálogo de Productos =====
        worksheet2 = workbook.create_sheet(title='Productos Disponibles')
        
        products = Product.objects.all().values('sku', 'description')
        df_products = pd.DataFrame(list(products))
        if not df_products.empty:
            df_products.columns = ['SKU', 'Descripción']
            # Escribir headers
            for col_idx, col_name in enumerate(df_products.columns, 1):
                cell = worksheet2.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
            # Escribir datos
            for row_idx, row_data in enumerate(df_products.values, 2):
                for col_idx, value in enumerate(row_data, 1):
                    worksheet2.cell(row=row_idx, column=col_idx, value=value)
        else:
            worksheet2.append(['SKU', 'Descripción'])
            for cell in worksheet2[1]:
                cell.fill = header_fill
                cell.font = header_font
        
        # Ajustar anchos
        for column in worksheet2.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet2.column_dimensions[column_letter].width = adjusted_width
        
        # ===== Sheet 3: Instrucciones =====
        worksheet3 = workbook.create_sheet(title='Instrucciones')
        instructions = [
            ['INSTRUCCIONES PARA IMPORTAR UNIDADES DE PRODUCTO'],
            [''],
            ['1. Complete los datos en la hoja "Unidades de Producto"'],
            ['2. SKU: Debe corresponder a un producto existente (ver hoja "Productos Disponibles")'],
            ['3. Nombre: Nombre de la unidad (ej: "Caja x12", "Pallet x48", "Pack x6")'],
            ['4. Conversion Factor: Factor de conversión a la unidad base del producto'],
            ['5. Campos obligatorios: SKU, Nombre, Conversion Factor'],
            [''],
            ['EJEMPLOS:'],
            ['• Si la unidad base es "unit" y crea "Caja x12", el factor es 12'],
            ['• Si la unidad base es "kg" y crea "Tonelada", el factor es 1000'],
            [''],
            ['IMPORTANTE:'],
            ['• Si ya existe una unidad con el mismo SKU y Nombre, se ACTUALIZARÁ'],
            ['• Si no existe, se CREARÁ una nueva unidad'],
            ['• NO modifique los nombres de las columnas'],
        ]
        
        for row in instructions:
            worksheet3.append(row)
        
        worksheet3.column_dimensions['A'].width = 80
        worksheet3['A1'].font = Font(bold=True, size=14)
        
        workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_unidades_producto.xlsx"'
        return response
    
    @action(detail=False, methods=['get'])
    def export(self, request):
        """Exportar todas las unidades de producto a CSV"""
        units = self.get_queryset().select_related('product')
        
        data = units.values(
            'id', 'name', 'conversion_factor', 
            'product__sku', 'product__description',
            'created_at', 'updated_at'
        )
        
        df = pd.DataFrame(list(data))
        
        df.columns = [
            'ID', 'Nombre', 'Conversion Factor',
            'SKU', 'Descripción Producto',
            'Fecha Creación', 'Fecha Actualización'
        ]
        
        # Formatear fechas
        if 'Fecha Creación' in df.columns:
            df['Fecha Creación'] = pd.to_datetime(df['Fecha Creación']).dt.strftime('%Y-%m-%d %H:%M:%S')
        if 'Fecha Actualización' in df.columns:
            df['Fecha Actualización'] = pd.to_datetime(df['Fecha Actualización']).dt.strftime('%Y-%m-%d %H:%M:%S')
        
        df = df.fillna('')
        
        output = BytesIO()
        df.to_csv(output, index=False, encoding='utf-8-sig', sep=';', decimal=',')
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='text/csv; charset=utf-8-sig'
        )
        response['Content-Disposition'] = 'attachment; filename="unidades_producto_export.csv"'
        return response
    
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser])
    def import_data(self, request):
        """Importar unidades de producto desde archivo Excel (crear/actualizar)"""
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No se proporcionó ningún archivo'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file = request.FILES['file']
        
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'El archivo debe ser un Excel (.xlsx o .xls)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            df = pd.read_excel(file, sheet_name='Unidades de Producto')
            
            column_mapping = {
                'SKU': 'sku',
                'Nombre': 'name',
                'Conversion Factor': 'conversion_factor'
            }
            
            df = df.rename(columns=column_mapping)
            
            # Validar columnas requeridas
            required_columns = ['sku', 'name', 'conversion_factor']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                reverse_mapping = {v: k for k, v in column_mapping.items()}
                missing_spanish = [reverse_mapping.get(col, col) for col in missing_columns]
                return Response(
                    {'error': f'Faltan columnas requeridas: {", ".join(missing_spanish)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # ===== OPTIMIZACIÓN: Precargar datos (2 queries) =====
            # 1. Precargar productos indexados por SKU (1 query)
            products_qs = Product.objects.all()
            products_map = {p.sku: p for p in products_qs}
            
            # 2. Precargar unidades existentes con clave compuesta (product_id, name) (1 query)
            existing_units_qs = ProductUnit.objects.select_related('product').all()
            existing_units_map = {(u.product.sku, u.name): u for u in existing_units_qs}
            
            units_to_create = []
            units_to_update = []
            errors = []
            
            # ===== Procesar datos en memoria (0 queries adicionales) =====
            for index, row in df.iterrows():
                try:
                    # Validar campos requeridos
                    if pd.isna(row['sku']) or pd.isna(row['name']) or pd.isna(row['conversion_factor']):
                        errors.append(f'Fila {index + 2}: SKU, Nombre y Conversion Factor son obligatorios')
                        continue
                    
                    sku = str(row['sku']).strip()
                    name = str(row['name']).strip()
                    
                    # Validar que el producto existe (lookup O(1))
                    product = products_map.get(sku)
                    if not product:
                        errors.append(f'Fila {index + 2}: Producto con SKU "{sku}" no existe')
                        continue
                    
                    # Validar conversion_factor
                    try:
                        conversion_factor = float(row['conversion_factor'])
                        if conversion_factor <= 0:
                            errors.append(f'Fila {index + 2}: Conversion Factor debe ser mayor a 0')
                            continue
                    except (ValueError, TypeError):
                        errors.append(f'Fila {index + 2}: Conversion Factor debe ser un número válido')
                        continue
                    
                    # Verificar si la unidad ya existe usando clave compuesta (lookup O(1))
                    existing_unit = existing_units_map.get((sku, name))
                    
                    # Preparar datos de la unidad
                    unit_data = {
                        'product': product,
                        'name': name,
                        'conversion_factor': conversion_factor
                    }
                    
                    if existing_unit:
                        # Actualizar unidad existente
                        for key, value in unit_data.items():
                            setattr(existing_unit, key, value)
                        units_to_update.append(existing_unit)
                    else:
                        # Crear nueva unidad
                        unit = ProductUnit(**unit_data)
                        units_to_create.append(unit)
                    
                except Exception as e:
                    errors.append(f'Fila {index + 2}: {str(e)}')
                    continue
            
            # ===== Ejecutar operaciones en bulk (2 queries totales) =====
            with transaction.atomic():
                # Crear unidades nuevas (1 query)
                if units_to_create:
                    ProductUnit.objects.bulk_create(units_to_create)
                
                # Actualizar unidades existentes (1 query)
                if units_to_update:
                    ProductUnit.objects.bulk_update(
                        units_to_update,
                        ['name', 'conversion_factor', 'product']
                    )
            
            return Response({
                'message': f'Importación completada. Creados: {len(units_to_create)}, Actualizados: {len(units_to_update)}',
                'created': len(units_to_create),
                'updated': len(units_to_update),
                'errors': errors if errors else []
            }, status=status.HTTP_200_OK if not errors else status.HTTP_207_MULTI_STATUS)
            
        except Exception as e:
            return Response(
                {'error': f'Error al procesar el archivo: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class StockViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet de solo lectura para Stock.
    El stock no puede ser creado, actualizado o eliminado directamente.
    Todas las modificaciones de stock deben realizarse a través de:
    - Órdenes de compra (PurchaseOrder)
    - Órdenes de venta (SalesOrder)
    - Movimientos de stock (StockMovement)
    """
    queryset = Stock.objects.select_related('product', 'warehouse', 'branch').all()
    serializer_class = StockSerializer
    permission_classes = [IsNotClientPermission]
    pagination_class = StockPagination
    
    def list(self, request, *args, **kwargs):
        """
        Listar todos los stocks con filtros opcionales y paginación.
        Query params:
        - product: ID del producto
        - warehouse: ID del depósito
        - branch: ID de la sucursal
        - low_stock: 'true' para mostrar solo productos con stock bajo
        - search: Buscar por nombre o SKU del producto
        - page: Número de página (default: 1)
        - page_size: Tamaño de página (default: 5, max: 100)
        """
        queryset = self.get_queryset()
        
        # Filtros opcionales
        product_id = request.query_params.get('product', None)
        warehouse_id = request.query_params.get('warehouse', None)
        branch_id = request.query_params.get('branch', None)
        low_stock = request.query_params.get('low_stock', None)
        search_term = request.query_params.get('search', None)
        
        if product_id:
            queryset = queryset.filter(product_id=product_id)
        
        if warehouse_id:
            queryset = queryset.filter(warehouse_id=warehouse_id)
        
        if branch_id:
            queryset = queryset.filter(branch_id=branch_id)
        
        # Filtrar por búsqueda de nombre o SKU del producto
        if search_term:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(product__description__icontains=search_term) |
                Q(product__sku__icontains=search_term) 
            )
        
        # Filtrar por stock bajo si se solicitó
        if low_stock == 'true':
            from django.db.models import F
            # Comparar quantity con el safety_stock del producto
            queryset = queryset.annotate(
                safety_stock=F('product__safety_stock')
            ).filter(quantity__lte=F('safety_stock'))
        
        # Aplicar paginación
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        # Si no hay paginación (no debería ocurrir), devolver todos
        serializer = self.get_serializer(queryset, many=True)
        return Response({
            'count': len(serializer.data),
            'results': serializer.data
        }, status=status.HTTP_200_OK)
    
    def retrieve(self, request, *args, **kwargs):
        """Obtener un registro de stock específico por ID"""
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'])
    def low_stock_alert(self, request):
        """
        Endpoint para obtener todos los productos con stock por debajo del nivel de seguridad.
        Útil para alertas y notificaciones.
        """
        queryset = self.get_queryset()
        
        low_stock_items = []
        for stock in queryset:
            if stock.product.safety_stock and stock.quantity < stock.product.safety_stock:
                low_stock_items.append({
                    'id': stock.id,
                    'product': {
                        'id': stock.product.id,
                        'sku': stock.product.sku,
                        'description': stock.product.description,
                        'safety_stock': float(stock.product.safety_stock),
                    },
                    'location_name': stock.warehouse.name if stock.warehouse else (stock.branch.name if stock.branch else 'Sin ubicación'),
                    'warehouse': stock.warehouse.name if stock.warehouse else None,
                    'branch': stock.branch.name if stock.branch else None,
                    'current_quantity': float(stock.quantity),
                    'difference': float(stock.product.safety_stock - stock.quantity),
                    'updated_at': stock.updated_at
                })
        
        return Response({
            'count': len(low_stock_items),
            'results': low_stock_items
        }, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'])
    def export_template(self, request):
        """Generar plantilla Excel para ajuste de stock"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        output = BytesIO()
        workbook = Workbook()
        
        # ===== Sheet 1: Stock Actual =====
        worksheet1 = workbook.active
        worksheet1.title = 'Ajuste de Stock'
        
        # Headers
        headers = ['ID Stock', 'SKU + Descripción', 'Ubicación', 'Cantidad Actual', 'Nueva Cantidad']
        worksheet1.append(headers)
        
        # Formato de headers
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF')
        for cell in worksheet1[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Obtener datos de stock actual
        stocks = self.get_queryset()
        for stock in stocks:
            sku_desc = f"{stock.product.sku} - {stock.product.description}"
            location = stock.warehouse.name if stock.warehouse else (stock.branch.name if stock.branch else 'Sin ubicación')
            
            worksheet1.append([
                stock.id,
                sku_desc,
                location,
                float(stock.quantity),
                '',  # Nueva cantidad (vacío para que el usuario complete)
            ])
        
        # Ajustar anchos de columna
        for column in worksheet1.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet1.column_dimensions[column_letter].width = adjusted_width
        
        # ===== Sheet 2: Instrucciones =====
        worksheet2 = workbook.create_sheet(title='Instrucciones')
        instructions = [
            ['INSTRUCCIONES PARA AJUSTE DE STOCK'],
            [''],
            ['1. Complete la columna "Nueva Cantidad" con la cantidad real contada'],
            ['2. Complete la columna "Motivo" con la razón del ajuste (ej: "Inventario físico", "Corrección de error", etc.)'],
            ['3. NO modifique las columnas: ID Stock, SKU + Descripción, Ubicación, Cantidad Actual'],
            ['4. Las filas que no tengan "Nueva Cantidad" serán ignoradas'],
            ['5. El sistema creará movimientos de ajuste automáticamente'],
            [''],
            ['IMPORTANTE:'],
            ['• Solo se procesarán las filas donde "Nueva Cantidad" sea diferente a "Cantidad Actual"'],
            ['• Los ajustes se registrarán como movimientos de tipo ADJUSTMENT'],
            ['• NO elimine ni agregue columnas a la hoja "Ajuste de Stock"'],
        ]
        
        for row in instructions:
            worksheet2.append(row)
        
        # Formato para instrucciones
        worksheet2.column_dimensions['A'].width = 80
        worksheet2['A1'].font = Font(bold=True, size=14)
        
        workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_ajuste_stock.xlsx"'
        return response
    
    @action(detail=False, methods=['get'])
    def export(self, request):
        """Exportar todos los stocks a CSV"""
        stocks = self.get_queryset()
        
        # Usar values() para obtener diccionarios directamente de la query
        data = stocks.values(
            'id', 'quantity', 'created_at', 'updated_at',
            'product__id', 'product__sku', 'product__description',
            'warehouse__id', 'warehouse__name',
            'branch__id', 'branch__name'
        )
        
        df = pd.DataFrame(list(data))
        
        # Renombrar columnas a español con formato especial para product
        df.columns = [
            'ID', 'Cantidad', 'Fecha Creación', 'Fecha Actualización',
            'ID Producto', 'SKU + Descripción', 'Descripción',
            'ID Depósito', 'Nombre Depósito',
            'ID Sucursal', 'Nombre Sucursal'
        ]
        
        # Combinar SKU + Descripción
        df['SKU + Descripción'] = df['SKU + Descripción'] + ' - ' + df['Descripción']
        df = df.drop(columns=['Descripción'])
       
        # Formatear fechas
        if 'Fecha Creación' in df.columns:
            df['Fecha Creación'] = pd.to_datetime(df['Fecha Creación']).dt.strftime('%Y-%m-%d %H:%M:%S')
        if 'Fecha Actualización' in df.columns:
            df['Fecha Actualización'] = pd.to_datetime(df['Fecha Actualización']).dt.strftime('%Y-%m-%d %H:%M:%S')
        
        df = df.fillna('')
        
        # Crear archivo CSV en memoria
        output = BytesIO()
        df.to_csv(output, index=False, encoding='utf-8-sig', sep=';', decimal=',')
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='text/csv; charset=utf-8-sig'
        )
        response['Content-Disposition'] = 'attachment; filename="stock_export.csv"'
        return response
    
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser])
    def import_data(self, request):
        """Importar ajustes de stock desde archivo Excel"""
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No se proporcionó ningún archivo'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file = request.FILES['file']
        
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'El archivo debe ser un Excel (.xlsx o .xls)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            df = pd.read_excel(file, sheet_name='Ajuste de Stock')
            
            column_mapping = {
                'ID Stock': 'stock_id',
                'Nueva Cantidad': 'new_quantity',
            }
            
            df = df.rename(columns=column_mapping)
            
            # ===== OPTIMIZACIÓN: Precargar registros de stock (1 query) =====
            # Obtener IDs únicos de stock del archivo
            stock_ids = df['stock_id'].dropna().unique().tolist()
            stock_ids = [int(sid) for sid in stock_ids if pd.notna(sid)]
            
            # Precargar stocks necesarios con select_related (1 query)
            stocks_qs = Stock.objects.filter(id__in=stock_ids).select_related('product', 'warehouse', 'branch')
            stocks_map = {s.id: s for s in stocks_qs}
            
            # Preparar listas para operaciones en bulk
            movements_to_create = []
            stocks_to_update = []
            adjustments_created = 0
            errors = []
            
            # ===== Procesar datos en memoria (0 queries adicionales) =====
            for index, row in df.iterrows():
                try:
                    # Saltar filas sin nueva cantidad
                    if pd.isna(row.get('new_quantity')) or row.get('new_quantity') == '':
                        continue
                    
                    # Validar ID Stock
                    stock_id = int(row['stock_id']) if pd.notna(row['stock_id']) else None
                    if not stock_id:
                        errors.append(f'Fila {index + 2}: ID Stock es obligatorio')
                        continue
                    
                    # Obtener registro de stock del diccionario (lookup O(1))
                    stock = stocks_map.get(stock_id)
                    if not stock:
                        errors.append(f'Fila {index + 2}: Stock con ID {stock_id} no existe')
                        continue
                    
                    # Validar nueva cantidad
                    try:
                        new_quantity = float(row['new_quantity'])
                    except (ValueError, TypeError):
                        errors.append(f'Fila {index + 2}: Nueva Cantidad debe ser un número')
                        continue
                    
                    # Calcular diferencia
                    current_quantity = float(stock.quantity)
                    difference = new_quantity - current_quantity
                    
                    # Solo procesar si hay diferencia
                    if abs(difference) < 0.0001:  # Evitar errores de precisión flotante
                        continue
                    
                    # Preparar movimiento de ajuste (se creará en bulk)
                    movement_type = 'IN' if difference > 0 else 'OUT'
                    quantity_abs = abs(difference)
                    
                    movement = StockMovement(
                        product=stock.product,
                        warehouse=stock.warehouse,
                        branch=stock.branch,
                        movement_type=movement_type,
                        quantity=quantity_abs,
                        status='REC',
                        date=timezone.now()
                    )
                    movements_to_create.append(movement)
                    
                    # Actualizar cantidad en memoria (se guardará en bulk)
                    stock.quantity = new_quantity
                    stocks_to_update.append(stock)
                    
                    adjustments_created += 1
                    
                except Exception as e:
                    errors.append(f'Fila {index + 2}: {str(e)}')
                    continue
            
            # ===== Ejecutar operaciones en bulk (2 queries totales) =====
            with transaction.atomic():
                # Crear movimientos de ajuste (1 query)
                if movements_to_create:
                    StockMovement.objects.bulk_create(movements_to_create)
                
                # Actualizar stocks (1 query)
                if stocks_to_update:
                    Stock.objects.bulk_update(stocks_to_update, ['quantity'])
            
            return Response({
                'message': f'Importación completada. Creados: {adjustments_created}, Actualizados: 0',
                'created': adjustments_created,
                'updated': 0,
                'errors': errors if errors else []
            }, status=status.HTTP_200_OK if not errors else status.HTTP_207_MULTI_STATUS)
            
        except Exception as e:
            return Response(
                {'error': f'Error al procesar el archivo: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class StockMovementViewSet(viewsets.ModelViewSet):
    """
    ViewSet para StockMovement.
    Los movimientos de stock pueden ser:
    - Consultados (GET)
    - Creados manualmente para transferencias internas (POST)
    Los movimientos de compras y ventas se crean automáticamente.
    """
    queryset = StockMovement.objects.select_related(
        'product', 
        'warehouse', 
        'branch', 
        'unit_used',
        'sale',
        'purchase'
    ).all().order_by('-date')
    serializer_class = StockMovementSerializer
    permission_classes = [IsNotClientPermission]
    pagination_class = StockMovementPagination
    http_method_names = ['get', 'post', 'patch', 'head', 'options']  # GET, POST y PATCH para actualizar estado
    
    def create(self, request, *args, **kwargs):
        """
        Crear un movimiento interno de stock.
        Valida que haya stock suficiente y actualiza las ubicaciones.
        """
        data = request.data
        
        # Validar campos requeridos
        required_fields = ['product', 'fromLocationType', 'fromLocation', 'toLocationType', 'toLocation', 'quantity']
        for field in required_fields:
            if field not in data or not data[field]:
                return Response(
                    {'error': f'El campo {field} es requerido'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        try:
            product = Product.objects.get(id=data['product'])
        except Product.DoesNotExist:
            return Response(
                {'error': 'Producto no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        from_type = data['fromLocationType']  # WHA o BRA
        from_id = data['fromLocation']
        to_type = data['toLocationType']  # WHA o BRA
        to_id = data['toLocation']
        quantity = Decimal(str(data['quantity']))
        note = data.get('note', '')
        
        # Validar que origen y destino sean diferentes
        if from_type == to_type and from_id == to_id:
            return Response(
                {'error': 'El origen y destino deben ser diferentes'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar cantidad
        if quantity <= 0:
            return Response(
                {'error': 'La cantidad debe ser mayor a 0'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Obtener ubicaciones
        from_warehouse = None
        from_branch = None
        to_warehouse = None
        to_branch = None
        
        try:
            if from_type == 'WHA':
                from_warehouse = Warehouse.objects.get(id=from_id)
            else:
                from_branch = Branch.objects.get(id=from_id)
                
            if to_type == 'WHA':
                to_warehouse = Warehouse.objects.get(id=to_id)
            else:
                to_branch = Branch.objects.get(id=to_id)
        except (Warehouse.DoesNotExist, Branch.DoesNotExist):
            return Response(
                {'error': 'Ubicación no encontrada'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Validar stock en origen
        try:
            if from_type == 'WHA':
                stock_origin = Stock.objects.get(product=product, warehouse=from_warehouse)
            else:
                stock_origin = Stock.objects.get(product=product, branch=from_branch)
                
            if stock_origin.quantity < quantity:
                location_name = from_warehouse.name if from_warehouse else from_branch.name
                return Response(
                    {
                        'error': f'Stock insuficiente en {location_name}',
                        'detail': f'Stock disponible: {stock_origin.quantity} {product.base_unit_name}, solicitado: {quantity} {product.base_unit_name}'
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
        except Stock.DoesNotExist:
            location_name = from_warehouse.name if from_warehouse else from_branch.name
            return Response(
                {'error': f'No hay stock de {product.description} en {location_name}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        with transaction.atomic():
            # Crear o actualizar stock en destino
            if to_type == 'WHA':
                stock_dest, created = Stock.objects.get_or_create(
                    product=product,
                    warehouse=to_warehouse,
                    defaults={'quantity': 0}
                )
            else:
                stock_dest, created = Stock.objects.get_or_create(
                    product=product,
                    branch=to_branch,
                    defaults={'quantity': 0}
                )
        
            # Actualizar cantidades
            stock_origin.quantity -= quantity
            stock_origin.save()
            
            stock_dest.quantity += quantity
            stock_dest.save()
            
            # Determinar qué ubicación usar para el movimiento (usamos la que sea warehouse o branch)
            movement_warehouse = from_warehouse or to_warehouse
            movement_branch = from_branch or to_branch
            
            # Preparar nota con metadata de ubicaciones para poder revertir el movimiento
            import json
            location_metadata = {
                'from_type': from_type,
                'from_id': int(from_id),  # Asegurar que sea int
                'to_type': to_type,
                'to_id': int(to_id)  # Asegurar que sea int
            }
            note_with_metadata = json.dumps(location_metadata) if not note else f"{json.dumps(location_metadata)}||{note}"
            
            # Crear el movimiento de stock
            # Los movimientos internos arrancan en tránsito (TRAN) por defecto
            # Solo se actualizan las cantidades cuando se marca como recibido (REC)
            # Por ahora actualizamos las cantidades inmediatamente pero el estado es TRAN
            movement = StockMovement.objects.create(
                product=product,
                warehouse=movement_warehouse,
                branch=movement_branch,
                from_location=from_type,
                to_location=to_type,
                movement_type='OUT' if from_type in ['WHA', 'BRA'] else 'IN',
                quantity=quantity,
                status='TRAN',  # Movimientos internos arrancan en tránsito
                note=note_with_metadata,
                conversion_factor_at_moment=Decimal('1.0')
            )
            
            # Agregar comentario al movimiento
            from_name = from_warehouse.name if from_warehouse else from_branch.name
            to_name = to_warehouse.name if to_warehouse else to_branch.name
            movement.add_comment(
                f'Movimiento interno: {quantity} {product.base_unit_name} de {from_name} a {to_name}',
                status_before=None,
                user=request.user
            )
            
            serializer = self.get_serializer(movement)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
    
    def list(self, request, *args, **kwargs):
        """
        Listar todos los movimientos de stock con filtros opcionales y paginación.
        Query params:
        - product: ID del producto
        - warehouse: ID del depósito
        - branch: ID de la sucursal
        - movement_type: Tipo de movimiento ('IN' o 'OUT')
        - status: Estado del movimiento ('PEN', 'TRAN', 'REC', 'CAN')
        - from_location: Origen del movimiento ('PUR', 'SAL', 'WHA', 'BRA', 'MOV')
        - to_location: Destino del movimiento ('PUR', 'SAL', 'WHA', 'BRA', 'MOV')
        - sale: ID de la orden de venta
        - purchase: ID de la orden de compra
        - date_from: Fecha desde (formato: YYYY-MM-DD)
        - date_to: Fecha hasta (formato: YYYY-MM-DD)
        - page: Número de página (default: 1)
        - page_size: Tamaño de página (default: 5, max: 100)
        """
        queryset = self.get_queryset()
        
        # Filtros opcionales
        product_id = request.query_params.get('product', None)
        warehouse_id = request.query_params.get('warehouse', None)
        branch_id = request.query_params.get('branch', None)
        movement_type = request.query_params.get('movement_type', None)
        status_filter = request.query_params.get('status', None)
        from_location = request.query_params.get('from_location', None)
        to_location = request.query_params.get('to_location', None)
        sale_id = request.query_params.get('sale', None)
        purchase_id = request.query_params.get('purchase', None)
        date_from = request.query_params.get('date_from', None)
        date_to = request.query_params.get('date_to', None)
        
        if product_id:
            queryset = queryset.filter(product_id=product_id)
        
        if warehouse_id:
            queryset = queryset.filter(warehouse_id=warehouse_id)
        
        if branch_id:
            queryset = queryset.filter(branch_id=branch_id)
        
        if movement_type and movement_type in ['IN', 'OUT']:
            queryset = queryset.filter(movement_type=movement_type)
        
        if status_filter and status_filter in ['PEN', 'TRAN', 'REC', 'CAN']:
            queryset = queryset.filter(status=status_filter)
        
        if from_location and from_location in ['PUR', 'SAL', 'WHA', 'BRA', 'MOV']:
            queryset = queryset.filter(from_location=from_location)
        
        if to_location and to_location in ['PUR', 'SAL', 'WHA', 'BRA', 'MOV']:
            queryset = queryset.filter(to_location=to_location)
        
        if sale_id:
            queryset = queryset.filter(sale_id=sale_id)
        
        if purchase_id:
            queryset = queryset.filter(purchase_id=purchase_id)
        
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        
        # Aplicar paginación
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        
        return Response({
            'count': len(serializer.data),
            'results': serializer.data
        }, status=status.HTTP_200_OK)
    
    def retrieve(self, request, *args, **kwargs):
        """Obtener un movimiento de stock específico por ID"""
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    def partial_update(self, request, *args, **kwargs):
        """
        Actualización parcial de un movimiento de stock.
        Solo permite actualizar el estado para movimientos internos 
        (aquellos que no están asociados a ventas ni compras).
        
        Reglas de transición de estados:
        - Una vez que está en 'REC' (Recibido), no se puede cambiar el estado
        - Si se cambia a 'CAN' (Cancelado), se revierten los cambios de stock
        """
        instance = self.get_object()
        
        # Verificar que sea un movimiento interno (sin venta ni compra asociada)
        if instance.sale is not None or instance.purchase is not None:
            return Response(
                {'error': 'Solo se puede modificar el estado de movimientos internos (sin venta ni compra asociada)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Solo permitir actualizar el campo status
        if 'status' not in request.data:
            return Response(
                {'error': 'Solo se puede actualizar el campo status'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        new_status = request.data.get('status')
        current_status = instance.status
        
        # Validar que el nuevo estado sea válido
        valid_statuses = ['PEN', 'TRAN', 'REC', 'CAN']
        if new_status not in valid_statuses:
            return Response(
                {'error': f'Estado inválido. Valores permitidos: {", ".join(valid_statuses)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # REGLA CRÍTICA: Una vez recibido o cancelado, no se puede cambiar el estado
        if current_status == 'REC':
            return Response(
                {'error': 'No se puede cambiar el estado de un movimiento ya recibido'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if current_status == 'CAN':
            return Response(
                {'error': 'No se puede cambiar el estado de un movimiento cancelado'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Si no hay cambio de estado, retornar sin hacer nada
        if current_status == new_status:
            return Response(self.get_serializer(instance).data, status=status.HTTP_200_OK)
        
        # Obtener ubicaciones para manejo de stock desde la metadata en note
        import json
        import logging
        logger = logging.getLogger(__name__)
        
        from_warehouse = None
        from_branch = None
        to_warehouse = None
        to_branch = None
        
        logger.info(f"=== Iniciando partial_update para movimiento {instance.id} ===")
        logger.info(f"Estado actual: {current_status}, Nuevo estado: {new_status}")
        logger.info(f"Note content: {instance.note}")
        logger.info(f"from_location: {instance.from_location}, to_location: {instance.to_location}")
        
        # Intentar parsear la metadata del note
        metadata_parsed = False
        try:
            if instance.note:
                # El note puede tener formato: "metadata||nota_usuario" o solo "metadata"
                note_parts = instance.note.split('||', 1)
                metadata_str = note_parts[0]
                logger.info(f"Intentando parsear metadata: {metadata_str}")
                
                location_metadata = json.loads(metadata_str)
                
                from_type = location_metadata['from_type']
                from_id = int(location_metadata['from_id'])  # Asegurar que sea int
                to_type = location_metadata['to_type']
                to_id = int(location_metadata['to_id'])  # Asegurar que sea int
                
                logger.info(f"Metadata parseada exitosamente: from={from_type}:{from_id}, to={to_type}:{to_id}")
                
                # Obtener las ubicaciones reales
                if from_type == 'WHA':
                    from_warehouse = Warehouse.objects.get(id=from_id)
                    logger.info(f"Warehouse origen encontrado: {from_warehouse.name} (ID: {from_warehouse.id})")
                elif from_type == 'BRA':
                    from_branch = Branch.objects.get(id=from_id)
                    logger.info(f"Branch origen encontrado: {from_branch.name} (ID: {from_branch.id})")
                    
                if to_type == 'WHA':
                    to_warehouse = Warehouse.objects.get(id=to_id)
                    logger.info(f"Warehouse destino encontrado: {to_warehouse.name} (ID: {to_warehouse.id})")
                elif to_type == 'BRA':
                    to_branch = Branch.objects.get(id=to_id)
                    logger.info(f"Branch destino encontrado: {to_branch.name} (ID: {to_branch.id})")
                
                metadata_parsed = True
        except (json.JSONDecodeError, KeyError, ValueError, Warehouse.DoesNotExist, Branch.DoesNotExist) as e:
            logger.error(f"Error al parsear metadata: {type(e).__name__}: {str(e)}")
            logger.info("NOTA: No se pudo parsear metadata del note. Este movimiento puede haber sido creado antes de la implementación de metadata.")
            
            # IMPORTANTE: Para movimientos antiguos sin metadata, NO podemos revertir el stock de forma segura
            # porque no sabemos las ubicaciones exactas de origen y destino
            if new_status == 'CAN':
                return Response(
                    {
                        'error': 'No se puede cancelar este movimiento porque no tiene información de ubicaciones. ' +
                                 'Solo los movimientos creados con la versión actual pueden ser cancelados. ' +
                                 f'Detalles técnicos: {str(e)}'
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        if new_status == 'CAN' and not metadata_parsed:
            return Response(
                {'error': 'No se puede cancelar este movimiento porque no se pudo obtener información de las ubicaciones.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Si se cambia a CANCELADO, revertir los cambios de stock
        if new_status == 'CAN':
            logger.info(f"Iniciando cancelación del movimiento {instance.id}")
            logger.info(f"from_warehouse: {from_warehouse}, from_branch: {from_branch}")
            logger.info(f"to_warehouse: {to_warehouse}, to_branch: {to_branch}")
            logger.info(f"Cantidad a revertir: {instance.quantity}")
            
            with transaction.atomic():
                try:
                    # Revertir: devolver stock al origen y quitarlo del destino
                    # ORIGEN: Devolver la cantidad que se quitó
                    if from_warehouse:
                        logger.info(f"Procesando reversión en warehouse origen: {from_warehouse.name}")
                        stock_origin, created = Stock.objects.get_or_create(
                            product=instance.product, 
                            warehouse=from_warehouse,
                            defaults={'quantity': 0}
                        )
                        cantidad_antes = stock_origin.quantity
                        stock_origin.quantity += instance.quantity  # Devolver al origen
                        stock_origin.save()
                        logger.info(f"Stock en origen: {cantidad_antes} -> {stock_origin.quantity}")
                    elif from_branch:
                        logger.info(f"Procesando reversión en branch origen: {from_branch.name}")
                        stock_origin, created = Stock.objects.get_or_create(
                            product=instance.product, 
                            branch=from_branch,
                            defaults={'quantity': 0}
                        )
                        cantidad_antes = stock_origin.quantity
                        stock_origin.quantity += instance.quantity
                        stock_origin.save()
                        logger.info(f"Stock en origen: {cantidad_antes} -> {stock_origin.quantity}")
                    else:
                        logger.warning("No se identificó ubicación de origen")
                    
                    # DESTINO: Quitar la cantidad que se agregó
                    if to_warehouse:
                        logger.info(f"Procesando reversión en warehouse destino: {to_warehouse.name}")
                        stock_dest = Stock.objects.get(product=instance.product, warehouse=to_warehouse)
                        cantidad_antes = stock_dest.quantity
                        stock_dest.quantity -= instance.quantity  # Quitar del destino
                        if stock_dest.quantity < 0:
                            stock_dest.quantity = 0  # Evitar stock negativo
                        stock_dest.save()
                        logger.info(f"Stock en destino: {cantidad_antes} -> {stock_dest.quantity}")
                    elif to_branch:
                        logger.info(f"Procesando reversión en branch destino: {to_branch.name}")
                        stock_dest = Stock.objects.get(product=instance.product, branch=to_branch)
                        cantidad_antes = stock_dest.quantity
                        stock_dest.quantity -= instance.quantity
                        if stock_dest.quantity < 0:
                            stock_dest.quantity = 0  # Evitar stock negativo
                        stock_dest.save()
                        logger.info(f"Stock en destino: {cantidad_antes} -> {stock_dest.quantity}")
                    else:
                        logger.warning("No se identificó ubicación de destino")
                    
                    # Actualizar el estado
                    instance.status = new_status
                    instance.save()
                    logger.info(f"Movimiento {instance.id} cancelado exitosamente")
                    
                    # Agregar comentario detallado
                    from_name = from_warehouse.name if from_warehouse else (from_branch.name if from_branch else 'desconocido')
                    to_name = to_warehouse.name if to_warehouse else (to_branch.name if to_branch else 'desconocido')
                    instance.add_comment(
                        f"Movimiento cancelado. Stock revertido: +{instance.quantity} a {from_name}, -{instance.quantity} de {to_name}",
                        status_before=current_status,
                        user=request.user
                    )
                    
                except Stock.DoesNotExist as e:
                    logger.error(f"Error Stock.DoesNotExist: {str(e)}")
                    return Response(
                        {'error': f'Error al revertir stock: No se encontró el registro de stock en destino. {str(e)}'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                except Exception as e:
                    logger.error(f"Error inesperado: {str(e)}")
                    return Response(
                        {'error': f'Error inesperado al revertir stock: {str(e)}'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
        else:
            # Para otros cambios de estado (PEN -> TRAN, TRAN -> REC, etc.)
            instance.status = new_status
            instance.save()
            
            # Agregar comentario sobre el cambio de estado
            status_labels = {
                'PEN': 'Pendiente de pago',
                'TRAN': 'En tránsito',
                'REC': 'Recibido',
                'CAN': 'Cancelado'
            }
            instance.add_comment(f"Estado actualizado de {status_labels.get(current_status, current_status)} a {status_labels.get(new_status, new_status)}")
        
        serializer = self.get_serializer(instance)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'])
    def by_product(self, request):
        """
        Obtener todos los movimientos de un producto específico.
        Query param requerido: product_id
        """
        product_id = request.query_params.get('product_id', None)
        
        if not product_id:
            return Response(
                {'error': 'Se requiere el parámetro product_id'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        queryset = self.get_queryset().filter(product_id=product_id)
        serializer = self.get_serializer(queryset, many=True)
        
        return Response({
            'product_id': product_id,
            'count': len(serializer.data),
            'results': serializer.data
        }, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'])
    def by_location(self, request):
        """
        Obtener todos los movimientos de una ubicación específica (warehouse o branch).
        Query params: warehouse_id o branch_id (uno de los dos requerido)
        """
        warehouse_id = request.query_params.get('warehouse_id', None)
        branch_id = request.query_params.get('branch_id', None)
        
        if not warehouse_id and not branch_id:
            return Response(
                {'error': 'Se requiere warehouse_id o branch_id'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        queryset = self.get_queryset()
        
        if warehouse_id:
            queryset = queryset.filter(warehouse_id=warehouse_id)
        elif branch_id:
            queryset = queryset.filter(branch_id=branch_id)
        
        serializer = self.get_serializer(queryset, many=True)
        
        return Response({
            'warehouse_id': warehouse_id,
            'branch_id': branch_id,
            'count': len(serializer.data),
            'results': serializer.data
        }, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'])
    def recent(self, request):
        """
        Obtener los movimientos más recientes.
        Query param: limit (default: 50, max: 100)
        """
        limit = int(request.query_params.get('limit', 50))
        limit = min(limit, 100)  # Máximo 100 registros
        
        queryset = self.get_queryset()[:limit]
        serializer = self.get_serializer(queryset, many=True)
        
        return Response({
            'limit': limit,
            'count': len(serializer.data),
            'results': serializer.data
        }, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'])
    def pending(self, request):
        """
        Obtener todos los movimientos pendientes (status='PEN' o 'TRAN').
        """
        queryset = self.get_queryset().filter(status__in=['PEN', 'TRAN'])
        serializer = self.get_serializer(queryset, many=True)
        
        return Response({
            'count': len(serializer.data),
            'results': serializer.data
        }, status=status.HTTP_200_OK)